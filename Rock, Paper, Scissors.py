import openpyxl
from random import randint
import os.path

def create_excel(file_path):
    headers = ["Game Number", "Player Choice", "Computer Choice", "Computer Result"]
    
    if not os.path.isfile(file_path):
        # Create a new Excel workbook
        wb = openpyxl.Workbook()
        sheet = wb.active
        
        # Add column headers
        sheet.append(headers)
        
        # Save the Excel workbook
        wb.save(file_path)
        wb.close()
    else:
        # Check if the main header row exists in the workbook
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        first_row = sheet[1]
        header_values = [cell.value for cell in first_row]

        if header_values != headers:
            # The main header row is missing or incorrect, recreate it
            sheet.insert_rows(1)
            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col_num)
                cell.value = header
        
            # Save the Excel workbook
            wb.save(file_path)
        
        wb.close()


def fill_excel(file_path, data):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    row_num = sheet.max_row + 1
    
    for row in data:
        if len(row) < 2:
            continue
        
        game_num = sheet.max_row - 1  # Game number
        computer_result = 2 if row[0] == row[1] else (1 if (row[1] == "Rock" and row[0] == "Scissors") or (row[0] == "Paper" and row[1] == "Scissors") or (row[0] == "Rock" and row[1] == "Paper") else 0)
        row_data = [game_num, row[0], row[1], computer_result]  # Data for the row to be saved
        
        sheet.append(row_data)  # Add the data row
        
    wb.save(file_path)
    wb.close()


g = ['Rock', 'Paper', 'Scissors']
player = True
win = 0
lose = 0
tie = 0

def play_again(player):
    new_game = input('Do you want to play again? Choose Y to play again - N to stop playing:')
    if new_game.lower() in ['y', 'yes']:
        print('\nOkay, let\'s play!!')
        player = True
    elif new_game.lower() in ['n', 'no']:
        print('\nOkay, see you next time.')
        player = False
    else:
        player = 2
    return player

file_path = "C:/Users/Admin/Desktop/DATA.xlsx" #MAKE SURE TO VERIFY THE PATH OF THE EXEL FILE 

# Create the Excel file with column headers
create_excel(file_path)

while player:
    data = []  # Reset the data list for each new game
    computer = g[randint(0, 2)]
    player_choice = input('\nChoose R for Rock - P for Paper - S for Scissors:').lower()
    player = player_choice in ['r', 'rock', 'p', 'paper', 's', 'scissors']

    if player:
        if player_choice == 'r':
            player_choice = 'Rock'
        elif player_choice == 'p':
            player_choice = 'Paper'
        elif player_choice == 's':
            player_choice = 'Scissors'

        print('\nYou chose:', player_choice)
        print('Computer chose:', computer)

        if player_choice == computer:
            print('TIE! It\'s a draw, you both chose', computer)
            tie += 1
        elif (player_choice == 'Rock' and computer == 'Scissors') or (player_choice == 'Paper' and computer == 'Rock') or (player_choice == 'Scissors' and computer == 'Paper'):
            print('You win!', player_choice, 'beats', computer)
            win += 1
        else:
            print('You lose!', computer, 'beats', player_choice)
            lose += 1

        data.append([player_choice, computer])
        fill_excel(file_path, data)
        print('Wins:', win, '/ Loses:', lose, '/ Ties:', tie)
        player = play_again(player)
#CREDIT RAMZIMIMOU